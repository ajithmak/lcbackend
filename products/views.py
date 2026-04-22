"""
products/views.py — PostgreSQL version.
BooleanField filters use True/False directly — standard Django ORM.
"""
import logging
import csv
import io

from rest_framework import generics, filters
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework import status as drf_status
from django.shortcuts import get_object_or_404
from django.utils.text import slugify

from core.mixins import SuccessResponseMixin, PaginatedResponseMixin
from core.exceptions import ProductNotFound

from .models import Product, Category
from .serializers import (
    ProductListSerializer,
    ProductDetailSerializer,
    CategorySerializer,
    StockUpdateSerializer,
)

logger = logging.getLogger(__name__)


def _parse_price(request, param):
    from rest_framework.exceptions import ValidationError
    raw = request.query_params.get(param)
    if raw is None:
        return None
    try:
        val = float(raw)
        if val < 0:
            raise ValueError
        return val
    except ValueError:
        raise ValidationError({param: f'"{raw}" is not a valid price.'})


# ── Category views ─────────────────────────────────────────────────────────────

class CategoryListView(PaginatedResponseMixin, generics.ListAPIView):
    serializer_class   = CategorySerializer
    permission_classes = [AllowAny]
    queryset           = Category.objects.filter(is_active=True)


class CategoryAdminListCreateView(
    SuccessResponseMixin, PaginatedResponseMixin, generics.ListCreateAPIView
):
    serializer_class   = CategorySerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset           = Category.objects.all().order_by('name')

    def create(self, request, *args, **kwargs):
        s = self.get_serializer(data=request.data)
        s.is_valid(raise_exception=True)
        cat = s.save()
        logger.info('Admin %s created category: %s', request.user.email, cat.name)
        return self.created(s.data, message=f'Category "{cat.name}" created.')


class CategoryAdminDetailView(
    SuccessResponseMixin, generics.RetrieveUpdateDestroyAPIView
):
    serializer_class   = CategorySerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset           = Category.objects.all()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        obj     = self.get_object()
        s       = self.get_serializer(obj, data=request.data, partial=partial)
        s.is_valid(raise_exception=True)
        cat = s.save()
        return self.ok(s.data, message=f'Category "{cat.name}" updated.')

    def destroy(self, request, *args, **kwargs):
        obj   = self.get_object()
        count = Product.objects.filter(category_id=obj.id, is_active=True).count()
        if count > 0:
            from rest_framework.exceptions import ValidationError
            raise ValidationError(
                f'Cannot delete "{obj.name}": {count} active product(s) assigned.'
            )
        name = obj.name
        obj.delete()
        logger.info('Admin %s deleted category: %s', request.user.email, name)
        return self.deleted(f'Category "{name}" deleted.')


# ── Public product views ────────────────────────────────────────────────────────

class ProductListView(PaginatedResponseMixin, generics.ListAPIView):
    """GET /api/v1/products/ — with filters"""
    serializer_class   = ProductListSerializer
    permission_classes = [AllowAny]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['name', 'description', 'tags']
    ordering_fields    = ['price', 'name', 'created_at', 'stock']
    ordering           = ['-created_at']

    def get_queryset(self):
        qs = Product.objects.filter(is_active=True)

        category_slug = self.request.query_params.get('category')
        if category_slug:
            try:
                cat = Category.objects.get(slug=category_slug)
                if not cat.is_active:
                    return Product.objects.none()
                qs = qs.filter(category_id=cat.id)
            except Category.DoesNotExist:
                return Product.objects.none()

        min_price = _parse_price(self.request, 'min_price')
        max_price = _parse_price(self.request, 'max_price')
        if min_price is not None and max_price is not None and min_price > max_price:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'min_price': 'min_price cannot exceed max_price.'})
        if min_price is not None:
            qs = qs.filter(price__gte=min_price)
        if max_price is not None:
            qs = qs.filter(price__lte=max_price)

        if self.request.query_params.get('featured') == 'true':
            qs = qs.filter(is_featured=True)
        if self.request.query_params.get('in_stock') == 'true':
            qs = qs.filter(stock__gt=0)

        return qs


class ProductDetailView(SuccessResponseMixin, APIView):
    """GET /api/v1/products/<slug>/"""
    permission_classes = [AllowAny]

    def get(self, request, slug):
        try:
            product = Product.objects.get(slug=slug)
            if not product.is_active:
                raise Product.DoesNotExist
        except Product.DoesNotExist:
            raise ProductNotFound(f'No product found with slug "{slug}".')
        return self.ok(ProductDetailSerializer(product, context={'request': request}).data)


class FeaturedProductsView(PaginatedResponseMixin, generics.ListAPIView):
    """GET /api/v1/products/featured/"""
    serializer_class   = ProductListSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return (
            Product.objects
            .filter(is_active=True, is_featured=True, stock__gt=0)
            .order_by('-created_at')
        )


# ── Admin product views ─────────────────────────────────────────────────────────

class AdminProductListCreateView(
    SuccessResponseMixin, PaginatedResponseMixin, generics.ListCreateAPIView
):
    serializer_class   = ProductDetailSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['name', 'description']
    ordering_fields    = ['price', 'stock', 'created_at', 'name']
    ordering           = ['-created_at']

    def get_queryset(self):
        qs = Product.objects.all()
        p  = self.request.query_params.get('is_active')
        if p == 'true':
            qs = qs.filter(is_active=True)
        elif p == 'false':
            qs = qs.filter(is_active=False)
        cat_name = self.request.query_params.get('category_name')
        if cat_name:
            cat_ids = list(
                Category.objects.filter(name__icontains=cat_name)
                .values_list('id', flat=True)
            )
            return qs.filter(category_id__in=cat_ids) if cat_ids else Product.objects.none()
        return qs

    def create(self, request, *args, **kwargs):
        s = self.get_serializer(data=request.data)
        s.is_valid(raise_exception=True)
        product = s.save()
        logger.info('Admin %s created product: %s', request.user.email, product.name)
        return self.created(s.data, message=f'Product "{product.name}" created.')


class AdminProductDetailView(
    SuccessResponseMixin, generics.RetrieveUpdateDestroyAPIView
):
    serializer_class   = ProductDetailSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = Product.objects.all()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        obj     = self.get_object()
        s       = self.get_serializer(obj, data=request.data, partial=partial)
        s.is_valid(raise_exception=True)
        product = s.save()
        logger.info('Admin %s updated product id=%s', request.user.email, product.id)
        return self.ok(s.data, message=f'Product "{product.name}" updated.')

    def destroy(self, request, *args, **kwargs):
        obj           = self.get_object()
        obj.is_active = False
        obj.save(update_fields=['is_active', 'updated_at'])
        logger.info('Admin %s deactivated product id=%s', request.user.email, obj.id)
        return self.deleted(f'Product "{obj.name}" deactivated.')


class AdminStockUpdateView(SuccessResponseMixin, APIView):
    """PATCH /api/v1/products/admin/<pk>/stock/"""
    permission_classes = [IsAuthenticated, IsAdminUser]

    def patch(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        s       = StockUpdateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        old           = product.stock
        product.stock = s.validated_data['stock']
        product.save(update_fields=['stock', 'updated_at'])
        logger.info('Admin %s stock id=%s: %s→%s', request.user.email, pk, old, product.stock)
        return self.ok(
            data={'id': product.id, 'name': product.name, 'stock': product.stock},
            message=f'Stock updated to {product.stock} units.',
        )


# ── CSV/XLSX Bulk Import ────────────────────────────────────────────────────────

def _to_bool(val):
    """Parse boolean from CSV string."""
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ('true', '1', 'yes')


def _safe_decimal(val, field):
    """Parse decimal, raise ValueError if invalid."""
    from decimal import Decimal, InvalidOperation
    try:
        return Decimal(str(val).strip())
    except InvalidOperation:
        raise ValueError(f'{field} must be a valid number (got "{val}")')


def _parse_csv_row(row, row_num, cat_map):
    """
    Validate and parse one CSV row into a dict ready for Product.objects.create().
    cat_map: {slug: Category, name_lower: Category}
    Returns (data_dict, error_string_or_None)
    """
    errors = []
    name = str(row.get('name', '')).strip()
    if not name or len(name) < 2:
        errors.append('name is required (min 2 chars)')

    # Category — match by slug or name (case-insensitive)
    cat_raw  = str(row.get('category', '')).strip()
    category = cat_map.get(cat_raw.lower()) or cat_map.get(slugify(cat_raw))
    if not category:
        errors.append(f'category "{cat_raw}" not found — use category slug or name')

    # Price
    price = None
    try:
        price = _safe_decimal(row.get('price', ''), 'price')
        if price <= 0:
            errors.append('price must be > 0')
    except ValueError as e:
        errors.append(str(e))

    # Original price (optional)
    original_price = None
    op_raw = str(row.get('original_price', '')).strip()
    if op_raw:
        try:
            original_price = _safe_decimal(op_raw, 'original_price')
            if price and original_price <= price:
                errors.append('original_price must exceed price')
        except ValueError as e:
            errors.append(str(e))

    # Stock
    stock = 0
    try:
        stock = int(str(row.get('stock', '0')).strip() or '0')
        if stock < 0:
            errors.append('stock cannot be negative')
    except ValueError:
        errors.append('stock must be a whole number')

    # Min order
    min_order = 1
    try:
        min_order = int(str(row.get('min_order', '1')).strip() or '1')
    except ValueError:
        pass

    if errors:
        return None, f'Row {row_num}: ' + '; '.join(errors)

    # Build unique slug
    base_slug = slugify(name)
    slug      = base_slug
    n         = 1
    while Product.objects.filter(slug=slug).exists():
        slug = f'{base_slug}-{n}'
        n += 1

    return {
        'name':           name,
        'slug':           slug,
        'description':    str(row.get('description', '')).strip(),
        'category':       category,
        'price':          price,
        'original_price': original_price,
        'stock':          stock,
        'min_order':      min_order,
        'unit_type':      str(row.get('unit_type', '')).strip().lower(),
        'image_url':      str(row.get('image_url', '')).strip(),
        'tags':           str(row.get('tags', '')).strip(),
        'is_featured':    _to_bool(row.get('is_featured', 'false')),
        'is_active':      _to_bool(row.get('is_active',   'true')),
    }, None


class AdminProductImportView(SuccessResponseMixin, APIView):
    """
    POST /api/v1/products/admin/import/
    Accept a CSV or XLSX file and bulk-create products.
    Returns summary: created count, skipped rows with reasons.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded. Send file as multipart field "file".'}, status=400)

        fname = file.name.lower()
        rows  = []

        # ── Parse file ──────────────────────────────────────────────────────
        if fname.endswith('.csv'):
            try:
                text   = file.read().decode('utf-8-sig')   # handles BOM
                reader = csv.DictReader(io.StringIO(text))
                rows   = list(reader)
            except Exception as e:
                return Response({'error': f'Could not parse CSV: {e}'}, status=400)

        elif fname.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb   = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws   = wb.active
                hdrs = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({hdrs[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
            except ImportError:
                return Response({'error': 'openpyxl is not installed. Use CSV format or contact admin.'}, status=400)
            except Exception as e:
                return Response({'error': f'Could not parse Excel file: {e}'}, status=400)
        else:
            return Response({'error': 'Unsupported file type. Upload a .csv or .xlsx file.'}, status=400)

        if not rows:
            return Response({'error': 'File is empty or has no data rows.'}, status=400)

        # ── Build category lookup map ────────────────────────────────────────
        all_cats = Category.objects.filter(is_active=True)
        cat_map  = {}
        for cat in all_cats:
            cat_map[cat.slug]            = cat
            cat_map[cat.name.lower()]    = cat
            cat_map[slugify(cat.name)]   = cat

        # ── Process rows ─────────────────────────────────────────────────────
        created  = []
        skipped  = []

        for i, row in enumerate(rows, start=2):   # row 1 = header
            # Skip completely blank rows
            if not any(str(v).strip() for v in row.values()):
                continue

            data, err = _parse_csv_row(row, i, cat_map)
            if err:
                skipped.append(err)
                continue

            try:
                product = Product.objects.create(**data)
                created.append(product.name)
            except Exception as e:
                skipped.append(f'Row {i} ("{data.get("name", "?")}): {e}')

        logger.info(
            'Admin %s bulk imported %d products (%d skipped)',
            request.user.email, len(created), len(skipped)
        )

        return self.ok(
            data={
                'created_count': len(created),
                'skipped_count': len(skipped),
                'created':       created,
                'skipped':       skipped,
            },
            message=f'{len(created)} product(s) imported successfully.'
                    + (f' {len(skipped)} row(s) skipped.' if skipped else ''),
        )
